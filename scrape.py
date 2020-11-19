import os
import sys
import csv
import tabula
import requests
import xlsxwriter
import pandas as pd
from time import sleep
from openpyxl import load_workbook
from bs4 import BeautifulSoup as Be
from anticaptchaofficial.recaptchav2proxyless import *


def solve_capture():
    solver = recaptchaV2Proxyless()
    solver.set_verbose(1)
    solver.set_key("67cf868b894c969c392310b070613ed9")
    solver.set_website_url("https://www.qkr.gov.al")
    solver.set_website_key("6LfcaioUAAAAAKe5mOhor1VqquuNuw_NyIbzdNUW")

    g_response = solver.solve_and_return_solution()
    if g_response != 0:
        return g_response
    else:
        print("task finished with error " + solver.error_code)


def search_key(Id):
    url = "http://www.qkr.gov.al/kerko/kerko-ne-regjistrin-tregtar/kerko-per-subjekt/"

    payload = {
        '__RequestVerificationToken': 'wgv0wCUbikBDY2bqIxOMALOM3qQLwxrH8dyGo3y84OZEOvuFu65gYRH3Dz8vroO1XTQ7IHqKsMTLx99_sd8rdSapCq1ALivWHfMhGQhjouM1',
        'Nipt': Id,
        'ufprt': '8354D8AD9CD23844D1F1C4FF15F16EB62C41B8DF181434A9E438CAF92CFB31C69E05A056C00BD3A3A3BE83182694363679724396E83D8ABAB63A08F9F687F228AFF5B57ACF8B915E1952A5AC46715EAB88F4E475B5D8AB6A059A8D7E0E5F7DE9FBC169ABE7FB11F028BF142C5C0AA2E43D2A9B05FA2EBCA53B8DF61182E07379BFC7BAC568930613D9568E4E07AB70D69244B5B766530DD3C32E3FB3189AEE3C'}

    headers = {
        'Cookie': 'ASP.NET_SessionId=4rz0fete2tyybhzvsv51d1hu; __RequestVerificationToken=Fb567JV4xM5XsxRHLZVHakArmnjwuLdlmKazC_-niEAlZ2ZDiLT2Rd8oD6iqkW4bljglZJYZXTBUr4T-s-KR5wbRTH_EUqkkyZ_Ka9oEuU81',
        'user-agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36"
    }

    try:
        response = requests.request("POST", url, headers=headers, data=payload)
        if response.status_code == requests.codes.ok:
            return response.text
        else:
            return None
    except ConnectionError:
        sleep(3)
        return search_key(Id)
    except Exception as e:
        print(e)
        return None


def parse(content):
    return Be(content, 'html5lib')


def write(lines):
    with open(file='result.csv', encoding='utf-8', mode='a', newline='') as csv_file:
        writer = csv.writer(csv_file, delimiter=',')
        writer.writerows(lines)


def write_empty_xlsx(filename):
    with xlsxwriter.Workbook(filename) as workbook:
        worksheet = workbook.add_worksheet()
        worksheet.write_row(0, 0, [])


def append_df_to_excel(lines, sheet_name='Sheet1', start_row=None, truncate_sheet=False, **to_excel_kwargs):
    filename = 'result.xlsx'
    if not os.path.isfile(filename):
        truncate_sheet = True
        write_empty_xlsx(filename)

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        writer.book = load_workbook(filename)

        if start_row is None and sheet_name in writer.book.sheetnames:
            start_row = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if start_row is None or truncate_sheet is True:
        start_row = 0

    df_head = pd.DataFrame(lines).T.iloc[[0]].style.set_properties(**{'background-color': '#D9D9D9'})
    df_body = pd.DataFrame(lines).T.iloc[[1]].style.set_properties(**{'background-color': '#38c276'})

    df_head.to_excel(writer, sheet_name, startrow=start_row, header=False, index=False, engine='openpyxl')
    df_body.to_excel(writer, sheet_name, startrow=start_row+1, header=False, index=False, engine='openpyxl')

    writer.save()


def get_indexes(rows):
    indexes = []
    for index, row in enumerate(rows):
        value = -1
        finalIndex = -1
        finalIndexes = []
        row_head = row[0].replace('.', '').strip()
        try:
            value = int(row_head[:1])
            finalIndex = index
            try:
                value = int(row_head[:2])
                finalIndex = index
                mainNum = row_head.strip()[:2]
                i = 0
                while True:
                    i = i + 1
                    subNum = mainNum + str(i)
                    # try:
                    #     int(rows[index+i+1][:1])
                    # except:
                    #     break
                    if subNum in row_head:
                        finalIndexes.append(index + i)
                        continue
                    break
                try:
                    value = int(row_head[:3])
                    finalIndex = index
                except:
                    pass
            except:
                pass
        except:
            pass
        if index > 0 and value == 1:
            break
        if finalIndex != -1 and finalIndex not in indexes:
            indexes.append(finalIndex)
        for i in finalIndexes:
            if i not in indexes:
                indexes.append(i)
    for i in range(indexes[-1], len(rows)):
        try:
            int(rows[i][0].strip()[:1])
            indexes.append(i+1)
            break
        except:
            continue
    return indexes


def get_head_num(head):
    value = -1
    row_head = head.replace('.', '').strip()
    try:
        value = int(row_head[:1])
        try:
            value = int(row_head[:2])
            mainNum = row_head.strip()[:2]
            i = 0
            while True:
                i = i + 1
                subNum = mainNum + str(i)
                if subNum in row_head:
                    finalIndexes.append(index + i)
                    continue
                break
            try:
                value = int(row_head[:3])
            except:
                pass
        except:
            pass
    except:
        pass
    return value


def check_first_column_is_full_number(column):
    try:
        int(column.replace('.', '').strip())
        return True
    except:
        return False


def check_rearrange(row):
    if check_first_column_is_full_number(row[0]):
        return True
    else:
        return False


def make_header(head_content, head_num):
    head_num = str(head_num)
    head_content_list = head_content.split(head_num)
    if len(head_content_list) > 1:
        head_content__sublist = head_content_list[1].split(head_num.split('.')[0])
        return head_num + ' ' + head_content__sublist[0]
    return head_content


def convert_to_csv(input_path):
    output = "output.csv"
    tabula.convert_into(input_path, output, output_format="csv", pages='all')
    with open(file=output, mode='r') as file:
        rows = list(csv.reader(file))
    # os.remove(output)
    # os.remove(input_path)
    indexes = get_indexes(rows)
    records = []
    head_content = ''
    temp_head_content = ''
    row_content = ''
    sub_dup_num = 0
    save_head_num = 0
    save_head = ''
    for key, index in enumerate(indexes[:-1]):
        start_index = index
        end_index = indexes[key+1]
        head_num = get_head_num(rows[start_index][0])
        for index_range in range(start_index, end_index):
            if index_range == start_index and head_num == -1:
                sub_dup_num += 1
                head_content = save_head
                head_num = str(save_head_num).split('.')[0] + '.' + str(sub_dup_num)
                for r in rows[index_range]:
                    row_content += r
            else:
                sub_dup_num = 0
                row = rows[index_range]
                if check_rearrange(row):
                    head_content += row[0]
                    head_content += ' '
                    head_content += row[1]
                    for r in row[2:]:
                        row_content += r
                else:
                    head_content += row[0]
                    for r in row[1:]:
                        row_content += r
        if head_content == save_head:
            temp_head_content = make_header(save_head, head_num)
        save_head_num = head_num
        save_head = head_content
        row_content = row_content.replace('\n', ' ')
        if temp_head_content != '':
            temp_head_content = temp_head_content.replace('\n', ' ')
            record = [temp_head_content, row_content]
        else:
            head_content = head_content.replace('\n', ' ')
            record = [head_content, row_content]
        head_content = ''
        row_content = ''
        temp_head_content = ''
        records.append(record)
    append_df_to_excel(lines=records)


def write_pdf(content):
    pdf_path = 'result.pdf'
    if os.path.isfile(pdf_path):
        os.remove(pdf_path)
    with open(file=pdf_path, mode='wb') as file:
        file.write(content)
    convert_to_csv(pdf_path)


def download(code):
    # url = "http://www.qkr.gov.al/umbraco/Surface/Documents/GenerateSimpleExtract"
    url = "http://www.qkr.gov.al/umbraco/Surface/Documents/GenerateHistoricExtract"

    recaptcha_response = solve_capture()

    payload = {'SubjectDefCode': code,
               'g-recaptcha-response': recaptcha_response
               }

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36',
        'Referer': 'http://www.qkr.gov.al/kerko/kerko-ne-regjistrin-tregtar/kerko-per-subjekt/',
        'Cookie': 'ASP.NET_SessionId=pxn25n4s1g2kzzdk1quko4ay; __RequestVerificationToken=06W4uFrgBbf2QndfrFWyyz19aSkYaoS0NumYit5ksMxT3OX30M6rJsMjC7lnhIgc9aXS53zoJNvSyAHAXvk5lxBnnZhO6OD7X7B0L225oC41; _ga=GA1.3.1010961615.1605535770; _gid=GA1.3.1222653889.1605535770; _gat=1'
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    write_pdf(response.content)


def read_txt():
    with open(file='20201111-DPT-TaxpayerList.csv', encoding='utf-8', mode='r') as file:
        rows = list(csv.reader(file))
    Ids = []
    for row in rows:
        Ids.append(row[0].split(';')[0].strip())
    return Ids


def main():
    Ids = sys.argv[1:]
    for Id in Ids:
        print(Id)
        response = search_key(Id)
        if response is not None:
            soup = parse(response)
            elements = soup.select(".result-element")
            for element in elements:
                subjectDefCode = element.find(id="subjectDefCode")['value']
                download(code=subjectDefCode)
                exit()


if __name__ == '__main__':
    main()
